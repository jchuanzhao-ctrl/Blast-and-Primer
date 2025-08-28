#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
位置找序列V14_染色体名称修复版.py
----------------------------------------------
修复了染色体名称映射问题：
1. 支持多种染色体名称格式（RefSeq, GenBank, UCSC）
2. 增强染色体名称映射逻辑
3. 改进错误消息
"""

import os
import re
import glob
import argparse
import tempfile
import subprocess
import traceback
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed

import pandas as pd
from openpyxl import Workbook
from Bio.Seq import Seq
from Bio.Blast import NCBIXML
import primer3
from pyfaidx import Fasta

# ---------- 默认配置 ----------
INPUT_XLSX = "off.xlsx"
OUTPUT_XLSX = "blast_to_position_results.xlsx"

UPSTREAM = 300
DOWNSTREAM = 300

PRIMER_PRODUCT_SIZE_MIN = 210
PRIMER_PRODUCT_SIZE_MAX = 250
PRIMER_TM_MIN = 54.0
PRIMER_TM_OPT = 57.0
PRIMER_TM_MAX = 60.0
PRIMER_TM_DIFF_MAX = 5.0
PRIMER_GC_MIN = 40.0
PRIMER_GC_MAX = 60.0
PRIMER_SIZE_MIN = 20
PRIMER_SIZE_OPT = 23
PRIMER_SIZE_MAX = 26

FWD_OFFSET_MIN = 70
FWD_OFFSET_MAX = 120
REV_PREF_MIN = 120
REV_PREF_MAX = 150

PRIMER_MAX_SELF_ANY = 8.0
PRIMER_MAX_HAIRPIN_TH = 35.0

BLAST_TIMEOUT = 240
LOCAL_MAX_PRIMER_CANDIDATES = int(os.getenv("LOCAL_MAX_PRIMER_CANDIDATES", "30"))
DEFAULT_FASTA_NAME = os.getenv("LOCAL_GENOME_FASTA", "GCA_000001405.28_GRCh38.p13_genomic.fna")
DEFAULT_DB_PREFIX = os.getenv("LOCAL_BLAST_DB_PREFIX", "GRCh38")

DNA_RE = re.compile(r'DNA\s*:\s*([A-Za-z]+)', re.IGNORECASE)

# 扩展染色体名称映射 - 包含带版本号和不带版本号的标识符
CHR_NAME_MAPPING = {
    # RefSeq 名称 (带版本号和不带版本号)
    "NC_000001.11": "1", "NC_000001": "1",
    "NC_000002.12": "2", "NC_000002": "2",
    "NC_000003.12": "3", "NC_000003": "3",
    "NC_000004.12": "4", "NC_000004": "4",
    "NC_000005.10": "5", "NC_000005": "5",
    "NC_000006.12": "6", "NC_000006": "6",
    "NC_000007.14": "7", "NC_000007": "7",
    "NC_000008.11": "8", "NC_000008": "8",
    "NC_000009.12": "9", "NC_000009": "9",
    "NC_000010.11": "10", "NC_000010": "10",
    "NC_000011.10": "11", "NC_000011": "11",
    "NC_000012.12": "12", "NC_000012": "12",
    "NC_000013.11": "13", "NC_000013": "13",
    "NC_000014.9": "14", "NC_000014": "14",
    "NC_000015.10": "15", "NC_000015": "15",
    "NC_000016.10": "16", "NC_000016": "16",
    "NC_000017.11": "17", "NC_000017": "17",
    "NC_000018.10": "18", "NC_000018": "18",
    "NC_000019.10": "19", "NC_000019": "19",
    "NC_000020.11": "20", "NC_000020": "20",
    "NC_000021.9": "21", "NC_000021": "21",
    "NC_000022.11": "22", "NC_000022": "22",
    "NC_000023.11": "X", "NC_000023": "X",
    "NC_000024.10": "Y", "NC_000024": "Y",
    "NC_012920.1": "MT", "NC_012920": "MT",
    
    # GenBank 名称 (带版本号和不带版本号)
    "CM000663.2": "1", "CM000663": "1",
    "CM000664.2": "2", "CM000664": "2",
    "CM000665.2": "3", "CM000665": "3",
    "CM000666.2": "4", "CM000666": "4",
    "CM000667.2": "5", "CM000667": "5",
    "CM000668.2": "6", "CM000668": "6",
    "CM000669.2": "7", "CM000669": "7",
    "CM000670.2": "8", "CM000670": "8",
    "CM000671.2": "9", "CM000671": "9",
    "CM000672.2": "10", "CM000672": "10",
    "CM000673.2": "11", "CM000673": "11",
    "CM000674.2": "12", "CM000674": "12",
    "CM000675.2": "13", "CM000675": "13",
    "CM000676.2": "14", "CM000676": "14",
    "CM000677.2": "15", "CM000677": "15",
    "CM000678.2": "16", "CM000678": "16",
    "CM000679.2": "17", "CM000679": "17",
    "CM000680.2": "18", "CM000680": "18",
    "CM000681.2": "19", "CM000681": "19",
    "CM000682.2": "20", "CM000682": "20",
    "CM000683.2": "21", "CM000683": "21",
    "CM000684.2": "22", "CM000684": "22",
    "CM000685.2": "X", "CM000685": "X",
    "CM000686.2": "Y", "CM000686": "Y",
    "J01415.2": "MT", "J01415": "MT",
    
    # UCSC 名称
    "chr1": "1", "chr2": "2", "chr3": "3", "chr4": "4",
    "chr5": "5", "chr6": "6", "chr7": "7", "chr8": "8",
    "chr9": "9", "chr10": "10", "chr11": "11", "chr12": "12",
    "chr13": "13", "chr14": "14", "chr15": "15", "chr16": "16",
    "chr17": "17", "chr18": "18", "chr19": "19", "chr20": "20",
    "chr21": "21", "chr22": "22", "chrX": "X", "chrY": "Y",
    "chrM": "MT"
}

# 反向映射
ACCESSION_TO_CHR = {k: v for k, v in CHR_NAME_MAPPING.items()}

BLAST_CACHE = {}
_FAIDX_CACHE = None
_ACC2KEY = None
_LABEL2KEY = None

# ---------------------- 常用工具 ----------------------
def detect_header_row(path, sheet_name, max_check=10):
    try:
        tmp = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
    except Exception:
        return None
    nr = min(len(tmp), max_check)
    for i in range(nr):
        row = tmp.iloc[i].astype(str).str.lower().fillna('')
        has_dna = any('dna' in v or 'sequence' in v or 'seq' in v for v in row)
        if has_dna:
            return i
    return None

def find_col(df, keywords):
    for c in df.columns:
        name = str(c).strip().lower()
        if any(k in name for k in keywords):
            return c
    return None

def extract_dna_from_row(df, row_idx, target_col, look_ahead):
    n = len(df)
    cols_to_check = [target_col] if target_col else df.columns
    for r_offset in range(look_ahead + 1):
        cur = row_idx + r_offset
        if cur >= n:
            break
        for c in cols_to_check:
            val = df.at[cur, c]
            if isinstance(val, str):
                m = DNA_RE.search(val)
                if m:
                    return m.group(1).upper()
                # 直接提取字母序列
                seq = ''.join(filter(str.isalpha, val)).upper()
                if 10 <= len(seq) <= 50:  # 更宽松的DNA序列长度范围
                    return seq
    return None

def mutation_positions(seq):
    if not seq:
        return []
    return [i + 1 for i, ch in enumerate(seq) if ch.islower()]

def parse_perfect_hits_from_records(blast_records, primer_len):
    perfect_hits = 0
    if not blast_records:
        return perfect_hits
    first = blast_records[0]
    if not getattr(first, "alignments", None):
        return perfect_hits
    for alignment in first.alignments:
        for hsp in getattr(alignment, "hsps", []):
            gaps = getattr(hsp, "gaps", 0) or 0
            if hsp.identities == primer_len and gaps == 0:
                perfect_hits += 1
    return perfect_hits

def score_pair(f_ok, r_ok, f_hits, r_hits, tm_f, tm_r):
    score = 0.0
    score += 2.0 if f_ok else 0.0
    score += 2.0 if r_ok else 0.0
    score -= min(f_hits or 0, 10)
    score -= min(r_hits or 0, 10)
    score -= abs(tm_f - tm_r) / 2.0
    score -= abs(tm_f - PRIMER_TM_OPT) / 3.0
    score -= abs(tm_r - PRIMER_TM_OPT) / 3.0
    return score

# ---------------------- Windows 静默子进程 ----------------------
def _silent_run(cmd, timeout=None, check=True, cwd=None):
    si = None
    cf = 0
    if os.name == 'nt':
        try:
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            cf |= getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
        except Exception:
            si = None
    return subprocess.run(cmd, timeout=timeout, check=check, cwd=cwd, startupinfo=si, creationflags=cf)

def _which(cmd):
    paths = os.environ.get("PATH", "").split(os.pathsep)
    pathext = os.environ.get("PATHEXT", ".EXE;.BAT;.CMD;.COM").split(os.pathsep)
    for p in paths:
        p = p.strip('"')
        full = os.path.join(p, cmd)
        if os.path.isfile(full):
            return full
        for ext in pathext:
            fp = full + ext
            if os.path.isfile(fp):
                return fp
    return None

def _find_exe(exe_name, base_dir):
    variants = [exe_name]
    if not exe_name.lower().endswith(".exe"):
        variants.append(exe_name + ".exe")
    for d in [base_dir, os.path.join(base_dir, "ncbi-blast-2.17.0+", "bin"), r"E:\blast", r"E:\blast\ncbi-blast-2.17.0+\bin"]:
        for v in variants:
            cand = os.path.join(d, v)
            if os.path.isfile(cand):
                return cand
    for v in variants:
        hit = _which(v)
        if hit:
            return hit
    return None

# ---------------------- 修复的数据库检查函数 ----------------------
def _db_files_exist(db_prefix_path):
    """检查BLAST数据库文件是否存在"""
    exts = [".nhr", ".nin", ".nsq", ".ndb", ".njs", ".not", ".00.nsq", ".00.nin", ".00.nhr", ".nal"]
    # 检查是否有任何数据库文件存在
    for ext in exts:
        if os.path.exists(f"{db_prefix_path}{ext}"):
            return True
    
    # 检查通配符匹配
    if glob.glob(f"{db_prefix_path}.*"):
        return True
    
    return False

# ---------------------- FASTA & 映射 ----------------------
def _ensure_faidx(fasta_path):
    global _FAIDX_CACHE
    if _FAIDX_CACHE is None:
        if not os.path.isfile(fasta_path):
            raise FileNotFoundError(f"未找到参考 FASTA：{fasta_path}")
        _FAIDX_CACHE = Fasta(fasta_path, as_raw=True, sequence_always_upper=False)
    return _FAIDX_CACHE

def _build_maps(fa):
    global _ACC2KEY, _LABEL2KEY
    if _ACC2KEY is not None and _LABEL2KEY is not None:
        return _ACC2KEY, _LABEL2KEY
    
    acc2key = {}
    label2key = {}
    
    # 匹配多种可能的标识符格式
    patterns = [
        r'\b([A-Z]{2}_[0-9]+(?:\.[0-9]+)?)\b',  # RefSeq: NC_000001.11
        r'\b(CM|CP|GCA)_[0-9]+(?:\.[0-9]+)?\b',  # GenBank: CM000663.2
        r'\b(chr[0-9XYM]+)\b',  # UCSC: chr1, chrX
        r'\b([0-9XYMT]+)\b'     # 纯数字/字母: 1, X, MT
    ]
    
    for key in fa.keys():
        try:
            ln = fa[key].long_name
        except Exception:
            ln = key
        
        # 提取所有可能的标识符
        identifiers = []
        for pattern in patterns:
            identifiers.extend(re.findall(pattern, ln, re.IGNORECASE))
        
        # 添加到映射 (带版本号和不带版本号)
        for identifier in identifiers:
            if isinstance(identifier, tuple):
                identifier = identifier[0]  # 处理匹配组
            
            # 添加完整标识符
            acc2key[identifier] = key
            acc2key[identifier.upper()] = key
            acc2key[identifier.lower()] = key
            
            # 添加基础标识符 (不带版本号)
            base_identifier = identifier.split('.')[0]
            acc2key[base_identifier] = key
            acc2key[base_identifier.upper()] = key
            acc2key[base_identifier.lower()] = key
        
        # 提取染色体标签
        low = ln.lower()
        m1 = re.search(r'chromosome\s+([0-9xy]+)', low)
        m2 = re.search(r'\bchr\s*([0-9xy]+)\b', low)
        m3 = re.search(r'\b([0-9XYMT]+)\b', low)
        
        label = None
        if m1:
            label = m1.group(1).upper()
        elif m2:
            label = m2.group(1).upper()
        elif m3:
            label = m3.group(1).upper()
        
        # 处理特殊染色体
        if not label:
            if 'mitochondria' in low or 'mt' in low:
                label = 'MT'
            elif 'x' in low:
                label = 'X'
            elif 'y' in low:
                label = 'Y'
        
        if label:
            label2key[label] = key
            label2key[label.upper()] = key
            label2key[label.lower()] = key
    
    _ACC2KEY, _LABEL2KEY = acc2key, label2key
    return _ACC2KEY, _LABEL2KEY

def fetch_seq_from_local(accession, start, end, base_dir, fasta_name=DEFAULT_FASTA_NAME):
    fa = _ensure_faidx(os.path.join(base_dir, fasta_name))
    acc2key, label2key = _build_maps(fa)
    
    # 尝试直接匹配
    if accession in fa.keys():
        key = accession
    elif accession in acc2key:
        key = acc2key[accession]
    elif accession in label2key:
        key = label2key[accession]
    else:
        # 尝试规范化名称
        normalized = accession.upper().replace("CHR", "")
        if normalized in acc2key:
            key = acc2key[normalized]
        elif normalized in label2key:
            key = label2key[normalized]
        else:
            # 尝试移除版本号
            base_acc = accession.split('.')[0]
            if base_acc in acc2key:
                key = acc2key[base_acc]
            elif base_acc in label2key:
                key = label2key[base_acc]
            else:
                # 最终尝试：使用所有可能的键
                possible_keys = [k for k in fa.keys() if accession in k]
                if possible_keys:
                    key = possible_keys[0]
                else:
                    raise KeyError(f"FASTA 中找不到序列：{accession}。可用键: {', '.join(fa.keys()[:5])}...")
    
    s = max(1, int(start))
    e = max(s, int(end))
    obj = fa.get_seq(key, s, e)
    seq = obj.seq if hasattr(obj, "seq") else str(obj)
    return seq

# ---------------------- BLAST 定位位置 (修复版) ----------------------
def _ensure_blastdb(base_dir, fasta_path, db_prefix):
    db_prefix_path = os.path.join(base_dir, db_prefix)
    if _db_files_exist(db_prefix_path):
        print(f"  使用现有BLAST数据库: {db_prefix_path}")
        return db_prefix_path
    
    print(f"  创建新的BLAST数据库: {db_prefix_path}")
    makeblastdb = _find_exe("makeblastdb", base_dir)
    if not makeblastdb:
        raise RuntimeError("未找到 makeblastdb。请确认 BLAST+ 安装并把 bin 放在脚本同目录或加入 PATH。")
    
    cmd = [makeblastdb, "-in", fasta_path, "-dbtype", "nucl", "-parse_seqids", "-out", db_prefix_path]
    print("  [makeblastdb] 正在构建本地数据库（一次性操作）...")
    try:
        result = _silent_run(cmd, timeout=3600)
        if result.returncode != 0:
            print(f"  [错误] makeblastdb 失败，退出码: {result.returncode}")
            print(f"  命令: {' '.join(cmd)}")
            raise RuntimeError("BLAST数据库创建失败")
        print("  [makeblastdb] 完成。")
        return db_prefix_path
    except Exception as e:
        print(f"  [错误] 创建BLAST数据库时出错: {e}")
        raise

def _auto_num_threads(override=None):
    if isinstance(override, int) and override >= 1:
        return override
    env = os.getenv("LOCAL_BLAST_NUM_THREADS")
    if env:
        try:
            n = int(env)
            if n >= 1:
                return n
        except Exception:
            pass
    cpu = os.cpu_count() or 4
    n = max(4, int(cpu * 0.75))
    return min(n, cpu)

def blast_locate_position(dna_short, base_dir, db_prefix, fasta_name, threads_override=None):
    """通过BLAST定位DNA序列在基因组中的位置（修复版）"""
    # 检查序列是否有效
    if not dna_short or len(dna_short) < 10:
        return None, None, None, None, f"序列太短({len(dna_short)}bp)，至少需要10bp"

    fasta_path = os.path.join(base_dir, fasta_name)
    try:
        db_prefix_path = _ensure_blastdb(base_dir, fasta_path, db_prefix)
    except Exception as e:
        return None, None, None, None, f"数据库错误: {str(e)}"

    blastn_exec = _find_exe("blastn", base_dir)
    if not blastn_exec:
        return None, None, None, None, "未找到 blastn 可执行文件"

    num_threads = _auto_num_threads(threads_override)

    q_path = o_path = None
    try:
        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".fasta") as qf:
            q_path = qf.name
            qf.write(f">query\n{dna_short}\n")

        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".xml") as of:
            o_path = of.name

        # 使用优化的BLAST参数
        cmd = [
            blastn_exec, 
            "-task", "blastn",  # 使用标准blastn任务
            "-query", q_path,
            "-db", db_prefix_path,
            "-outfmt", "5",
            "-evalue", "10",  # 宽松的e值阈值
            "-word_size", "11",  # 中等字大小
            "-perc_identity", "90",  # 最低90%相似度
            "-num_alignments", "3",  # 获取多个匹配
            "-num_threads", str(num_threads),
            "-out", o_path
        ]
        
        try:
            _silent_run(cmd, timeout=BLAST_TIMEOUT)
        except Exception as e:
            return None, None, None, None, f"BLAST执行错误: {str(e)}"

        try:
            with open(o_path, "r") as h:
                records = list(NCBIXML.parse(h))
        except Exception as e:
            return None, None, None, None, f"BLAST XML解析错误: {str(e)}"

        if not records or not records[0].alignments:
            return None, None, None, None, "BLAST未找到匹配"

        # 获取最佳匹配
        best_alignment = records[0].alignments[0]
        hsp = best_alignment.hsps[0]
        
        # 确定链方向
        strand = '+' if hsp.sbjct_start < hsp.sbjct_end else '-'
        
        # 计算中心位置
        if strand == '+':
            center_pos = hsp.sbjct_start + len(dna_short) // 2
        else:
            center_pos = hsp.sbjct_end - len(dna_short) // 2
        
        # 获取染色体标识符
        chr_acc = best_alignment.accession
        
        # 转换为标准染色体名称
        standard_chr = CHR_NAME_MAPPING.get(chr_acc, chr_acc)
        if '.' in chr_acc:  # 尝试不带版本号
            base_acc = chr_acc.split('.')[0]
            standard_chr = CHR_NAME_MAPPING.get(base_acc, standard_chr)
        
        return standard_chr, center_pos, strand, chr_acc, "成功定位"

    except Exception as e:
        return None, None, None, None, f"BLAST错误: {str(e)}"
    finally:
        for p in [q_path, o_path]:
            try:
                if p and os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass

# ---------------------- BLAST 验证特异性 ----------------------
def _blast_batch(primer_list, base_dir, db_prefix, fasta_name, threads_override=None):
    pending = [p for p in primer_list if p not in BLAST_CACHE]
    results = {p: BLAST_CACHE[p] for p in primer_list if p in BLAST_CACHE}
    if not pending:
        return results

    fasta_path = os.path.join(base_dir, fasta_name)
    try:
        db_prefix_path = _ensure_blastdb(base_dir, fasta_path, db_prefix)
    except Exception as e:
        print(f"  [错误] 无法准备BLAST数据库: {e}")
        # 为所有待处理序列返回错误
        for seq in pending:
            results[seq] = (False, f"数据库错误: {e}", 1000)
        return results

    blastn_exec = _find_exe("blastn", base_dir)
    if not blastn_exec:
        # 为所有待处理序列返回错误
        for seq in pending:
            results[seq] = (False, "未找到blastn", 1000)
        return results

    num_threads = _auto_num_threads(threads_override)

    q_path = o_path = None
    try:
        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".fasta") as qf:
            q_path = qf.name
            for i, seq in enumerate(pending):
                qf.write(f">q{i}\n{seq}\n")
        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".xml") as of:
            o_path = of.name

        cmd = [
            blastn_exec, "-task", "blastn-short",
            "-query", q_path,
            "-db", db_prefix_path,
            "-outfmt", "5",
            "-evalue", "1000",
            "-num_threads", str(num_threads),
            "-out", o_path
        ]
        _silent_run(cmd, timeout=BLAST_TIMEOUT)

        with open(o_path, "r") as h:
            records = list(NCBIXML.parse(h))

        for i, rec in enumerate(records):
            if i >= len(pending):
                break
            seq = pending[i]
            L = len(seq)
            perfect = parse_perfect_hits_from_records([rec], L)
            try:
                top = rec.alignments[0].hsps[0]
                ident_ratio = (top.identities or 0) / max(1, L)
            except Exception:
                ident_ratio = 0.0
            if perfect == 0:
                res = (True, f"Specific (no perfect hits, best: {ident_ratio:.1%})", 0)
            elif perfect == 1:
                res = (True, "Specific (1 perfect hit)", 1)
            elif 1 < perfect <= 5:
                res = (True, f"Likely specific ({perfect} perfect hits)", perfect)
            else:
                res = (False, f"Non-specific ({perfect} perfect hits)", perfect)
            BLAST_CACHE[seq] = res
            results[seq] = res

        for seq in pending[len(records):]:
            res = (False, "BLAST missing", 10 ** 6)
            BLAST_CACHE[seq] = res
            results[seq] = res
        return results
    except Exception as e:
        print(f"  [错误] BLAST特异性检查失败: {e}")
        # 为所有待处理序列返回错误
        for seq in pending:
            results[seq] = (False, f"BLAST错误: {str(e)}", 1000)
        return results
    finally:
        for p in [q_path, o_path]:
            try:
                if p and os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass

# ---------------------- 设计引物 ----------------------
def design_and_check_primers(seq600, target_center, target_len, strand='+',
                             base_dir=None, db_prefix=DEFAULT_DB_PREFIX, fasta_name=DEFAULT_FASTA_NAME,
                             threads_override=None):
    if not seq600 or len(seq600) < PRIMER_PRODUCT_SIZE_MIN:
        return [], []

    if base_dir is None:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    def design_once(level):
        fmin, fmax = FWD_OFFSET_MIN, FWD_OFFSET_MAX
        rmin, rmax = REV_PREF_MIN, REV_PREF_MAX
        pmin, pmax = PRIMER_PRODUCT_SIZE_MIN, PRIMER_PRODUCT_SIZE_MAX
        tmin, topt, tmax = PRIMER_TM_MIN, PRIMER_TM_OPT, PRIMER_TM_MAX
        gcmin, gcmax = PRIMER_GC_MIN, PRIMER_GC_MAX
        lmin, lopt, lmax = PRIMER_SIZE_MIN, PRIMER_SIZE_OPT, PRIMER_SIZE_MAX
        self_any, hairpin = PRIMER_MAX_SELF_ANY, PRIMER_MAX_HAIRPIN_TH
        num_return = {0: LOCAL_MAX_PRIMER_CANDIDATES, 1: LOCAL_MAX_PRIMER_CANDIDATES,
                      2: max(LOCAL_MAX_PRIMER_CANDIDATES, 50), 3: max(LOCAL_MAX_PRIMER_CANDIDATES, 60),
                      4: max(LOCAL_MAX_PRIMER_CANDIDATES, 70), 5: max(LOCAL_MAX_PRIMER_CANDIDATES, 80)}.get(level, LOCAL_MAX_PRIMER_CANDIDATES)

        if level >= 1:
            fmin, fmax = 60, 130
            rmin, rmax = REV_PREF_MIN - 20, REV_PREF_MAX + 20
        if level >= 2:
            pmin, pmax = 200, 260
        if level >= 3:
            tmin, topt, tmax = 53.0, 57.0, 62.0
            gcmin, gcmax = 35.0, 65.0
            lmin, lopt, lmax = 19, 23, 27
        if level >= 4:
            self_any, hairpin = 10.0, 45.0
        if level >= 5:
            pmin, pmax = 190, 260
            fmin, fmax = 55, 140

        revcomp_output = False
        tc = target_center
        seq_template = seq600
        if strand == '-':
            seq_template = str(Seq(seq_template).reverse_complement())
            tc = len(seq_template) - tc - 1
            revcomp_output = True

        ts = tc - (target_len // 2)
        seq_args = {'SEQUENCE_ID': 'target_sequence', 'SEQUENCE_TEMPLATE': seq_template,
                    'SEQUENCE_TARGET': [ts, target_len]}
        global_args = {
            'PRIMER_TASK': 'generic', 'PRIMER_PICK_LEFT_PRIMER': 1, 'PRIMER_PICK_RIGHT_PRIMER': 1,
            'PRIMER_NUM_RETURN': int(num_return),
            'PRIMER_PRODUCT_SIZE_RANGE': [int(pmin), int(pmax)],
            'PRIMER_MIN_SIZE': int(lmin), 'PRIMER_OPT_SIZE': int(lopt), 'PRIMER_MAX_SIZE': int(lmax),
            'PRIMER_MIN_TM': float(tmin), 'PRIMER_OPT_TM': float(topt), 'PRIMER_MAX_TM': float(tmax),
            'PRIMER_PAIR_MAX_DIFF_TM': float(PRIMER_TM_DIFF_MAX),
            'PRIMER_MIN_GC': float(gcmin), 'PRIMER_MAX_GC': float(gcmax),
            'PRIMER_MAX_SELF_ANY': float(self_any), 'PRIMER_MAX_HAIRPIN_TH': float(hairpin)
        }

        try:
            p3 = primer3.design_primers(seq_args, global_args)
        except Exception as e:
            print(f"  Primer3 设计失败(L{level}): {e}")
            return [], []

        num_pairs = p3.get('PRIMER_PAIR_NUM_RETURNED', 0)
        if num_pairs == 0:
            print(f"  L{level}: Primer3 未找到候选")
            return [], []

        candidates = []
        for i in range(min(num_pairs, num_return)):
            f_seq = p3[f'PRIMER_LEFT_{i}_SEQUENCE']
            r_seq = p3[f'PRIMER_RIGHT_{i}_SEQUENCE']

            if revcomp_output:
                f_seq = str(Seq(f_seq).reverse_complement())
                r_seq = str(Seq(r_seq).reverse_complement())

            f_pos = p3[f'PRIMER_LEFT_{i}'][0]
            r_pos = p3[f'PRIMER_RIGHT_{i}'][0]

            try:
                amplicon = seq_template[f_pos: r_pos + 1]
                if revcomp_output:
                    amplicon = str(Seq(amplicon).reverse_complement())
                amp_size = len(amplicon)
            except Exception:
                amplicon, amp_size = "", 0

            f_offset = tc - f_pos
            r_offset = r_pos - tc

            f_in_range = (fmin <= f_offset <= fmax)
            prod_ok = (pmin <= amp_size <= pmax)
            r_pref_ok = (rmin <= r_offset <= rmax)

            tm_f = round(p3[f'PRIMER_LEFT_{i}_TM'], 2)
            tm_r = round(p3[f'PRIMER_RIGHT_{i}_TM'], 2)

            candidates.append({
                "f_seq": f_seq, "r_seq": r_seq, "amplicon": amplicon, "amp_size": amp_size,
                "tm_f": tm_f, "tm_r": tm_r,
                "f_offset": f_offset, "r_offset": r_offset,
                "f_in_range": f_in_range, "prod_ok": prod_ok, "r_pref_ok": r_pref_ok
            })

        mid_r = (rmin + rmax) / 2.0
        candidates.sort(key=lambda c: (int(c["f_in_range"]), int(c["prod_ok"]), int(c["r_pref_ok"]), -abs(c["r_offset"] - mid_r)), 
                         reverse=True)

        seqs = []
        for c in candidates:
            seqs.append(c["f_seq"])
            seqs.append(c["r_seq"])
        spec_map = _blast_batch(seqs, base_dir=base_dir, db_prefix=db_prefix, fasta_name=fasta_name,
                               threads_override=threads_override)

        best_perfect = None
        best_any = None
        best_any_score = -1e9
        for c in candidates:
            f = c["f_seq"]
            r = c["r_seq"]
            f_ok, f_msg, f_hits = spec_map.get(f, (False, "BLAST missing", 10 ** 6))
            r_ok, r_msg, r_hits = spec_map.get(r, (False, "BLAST missing", 10 ** 6))
            pair_score = score_pair(f_ok, r_ok, f_hits, r_hits, c["tm_f"], c["tm_r"])
            perfect = c["f_in_range"] and c["prod_ok"] and f_ok and r_ok

            if perfect and (best_perfect is None or pair_score > best_perfect["pair_score"]):
                best_perfect = {**c, "f_msg": f_msg, "r_msg": r_msg, "pair_score": pair_score}
            if pair_score > best_any_score:
                best_any = {**c, "f_msg": f_msg, "r_msg": r_msg, "pair_score": pair_score}
                best_any_score = pair_score

        if best_perfect:
            bp = best_perfect
            return [(
                bp["f_seq"], bp["r_seq"], bp["amplicon"],
                bp["tm_f"], bp["tm_r"],
                bp["f_msg"], bp["r_msg"],
                "Perfect", False,
                bp["f_offset"], bp["r_offset"], bp["r_pref_ok"], bp["amp_size"]
            )], candidates

        if best_any:
            ba = best_any
            return [(
                ba["f_seq"], ba["r_seq"], ba["amplicon"],
                ba["tm_f"], ba["tm_r"],
                ba["f_msg"], ba["r_msg"],
                "Suboptimal", True,
                ba["f_offset"], ba["r_offset"], ba["r_pref_ok"], ba["amp_size"]
            )], candidates

        return [], []

    last_cands = []
    ret_any = []
    for L in range(0, 6):
        print(f"  [Design] 放宽级别 L{L}...")
        ret, cands = design_once(L)
        if cands:
            last_cands = cands
        if ret:
            if ret[0][7] == "Perfect":
                return ret, last_cands
            ret_any = ret
    return ret_any, last_cands

# ---------------------- Worker ----------------------
def worker_process(task):
    """
    单条位点的完整处理。task 包含：
      sheet, dna_short, base_dir, fasta_name, db_prefix, blast_threads
    返回 (sheet, output_row)
    """
    (sheet, dna_short, base_dir, fasta_name, db_prefix, blast_threads) = task

    try:
        # 1. 通过BLAST定位位置
        if not dna_short:
            return sheet, ["", "", "", dna_short or "", "", "ERROR: DNA sequence is empty",
                           "No primers", "No primers", "",
                           "", "", "", "", "", "", "", "",
                           "", "", ""]

        # 记录原始序列
        original_dna = dna_short
        
        # 清理序列，只保留字母字符
        dna_short = ''.join(filter(str.isalpha, dna_short)).upper()
        
        if len(dna_short) < 10:
            return sheet, ["", "", "", original_dna, "", f"ERROR: 序列太短({len(dna_short)}bp)，至少需要10bp",
                           "No primers", "No primers", "",
                           "", "", "", "", "", "", "", "",
                           "", "", ""]

        # 定位位置
        print(f"  定位序列: {dna_short[:20]}... (长度: {len(dna_short)}bp)")
        chrom, posn, strand, chr_acc, blast_status = blast_locate_position(
            dna_short, base_dir, db_prefix, fasta_name, blast_threads
        )

        if not chrom or not posn:
            return sheet, ["", "", "", original_dna, "", f"ERROR: BLAST定位失败 - {blast_status}",
                           "No primers", "No primers", "",
                           "", "", "", "", "", "", "", "",
                           "", "", ""]

        print(f"  定位成功: 染色体 {chrom}, 位置 {posn}, 链 {strand}")

        # 2. 提取序列
        try:
            start_pos = max(1, posn - UPSTREAM)
            end_pos = posn + DOWNSTREAM
            print(f"  提取序列: {chr_acc}:{start_pos}-{end_pos}")
            seq600 = fetch_seq_from_local(chr_acc, start_pos, end_pos,
                                         base_dir=base_dir, fasta_name=fasta_name)
            if len(seq600) < 600:
                return sheet, [chrom, posn, strand, original_dna, "",
                               f"WARNING: 提取序列长度不足({len(seq600)}bp)", "No primers", "No primers", "",
                               "", "", "", "", "", "", "", "",
                               "", "", ""]
        except Exception as e:
            return sheet, [chrom, posn, strand, original_dna, "",
                           f"ERROR: FASTA提取失败 - {e}", "No primers", "No primers", "",
                           "", "", "", "", "", "", "", "",
                           "", "", ""]

        # 3. 设计引物
        print("  设计引物...")
        pairs, _cand = design_and_check_primers(
            seq600, UPSTREAM, len(dna_short), strand,
            base_dir=base_dir, db_prefix=db_prefix, fasta_name=fasta_name,
            threads_override=blast_threads
        )

        if not pairs:
            return sheet, [chrom, posn, strand, original_dna, "",
                           seq600, "No primers", "No primers", "",
                           "", "", "", "", "", "", "", "",
                           "", "", ""]

        (f_primer, r_primer, amp_seq, tm_f, tm_r,
         f_spec, r_spec, status, non_perfect,
         f_offset, r_offset, r_pref_ok, amp_size) = pairs[0]

        mutpos = mutation_positions(dna_short or "")

        return sheet, [
            chrom, posn, strand,
            original_dna,
            ", ".join(map(str, mutpos)) if mutpos else "",
            seq600,
            "Pair 1", status, "TRUE" if non_perfect else "FALSE",
            f_primer, tm_f, f_spec,
            r_primer, tm_r, r_spec,
            amp_seq, amp_size,
            f_offset if f_offset is not None else "",
            r_offset if r_offset is not None else "",
            "TRUE" if r_pref_ok else "FALSE"
        ]

    except Exception as e:
        traceback.print_exc()
        return sheet, ["", "", "", dna_short or "", "",
                       f"ERROR: {str(e)}", "No primers", "No primers", "",
                       "", "", "", "", "", "", "", "",
                       "", "", ""]

# ---- 辅助：用于探测进程池是否能正常启动（WinError 1455 早发现） ----
def _noop(x):
    return x

def _try_make_process_pool(n_workers):
    """尝试创建进程池并提交一个最小任务；若失败抛出异常。"""
    ex = ProcessPoolExecutor(max_workers=n_workers)
    fut = ex.submit(_noop, 1)  # 顶层函数，可 picklable
    fut.result(timeout=5)
    return ex

# ---------------------- 主流程 ----------------------
def main():
    parser = argparse.ArgumentParser(description="引物设计 v14（染色体名称修复版）")
    parser.add_argument("-i", "--input", default=INPUT_XLSX, help="输入 Excel 文件（默认 seq.xlsx）")
    parser.add_argument("-o", "--output", default=OUTPUT_XLSX, help="输出 Excel 文件名")
    parser.add_argument("--basedir", default=None, help="基准目录（默认=脚本所在目录，例如 E:\\blast）")
    parser.add_argument("--fasta", default=DEFAULT_FASTA_NAME, help="参考 FASTA 文件名或路径（默认环境变量或同目录）")
    parser.add_argument("--dbprefix", default=DEFAULT_DB_PREFIX, help="本地 BLAST 数据库前缀（默认 GRCh38）")
    parser.add_argument("--jobs", type=int, default=0, help="并行进程/线程数（0=自动=CPU//2，至少2）")
    parser.add_argument("--blast_threads", type=int, default=0, help="每个任务内部 BLAST 的线程数（0=自动分配）")
    parser.add_argument("--chunk", type=int, default=16, help="分批提交的任务块大小（默认32，避免一次性提交过多）")
    parser.add_argument("--force_executor", choices=["auto", "process", "thread"], default="auto", help="优先使用的执行器类型（默认 auto）")
    args = parser.parse_args()

    base_dir = args.basedir or os.path.dirname(os.path.abspath(__file__))
    fasta_name = args.fasta
    db_prefix = args.dbprefix
    input_xlsx = args.input
    output_xlsx = args.output

    if not os.path.exists(input_xlsx):
        print(f"找不到输入文件：{input_xlsx}")
        return

    # 预先构建 FASTA 索引与 BLAST 数据库（避免并发冲突）
    try:
        print("准备参考数据...")
        fa = _ensure_faidx(os.path.join(base_dir, fasta_name))
        _build_maps(fa)
        _ensure_blastdb(base_dir, os.path.join(base_dir, fasta_name), db_prefix)
        print("参考数据准备完成")
        print(f"FASTA 序列键示例: {list(fa.keys())[:5]}")
    except Exception as e:
        print(f"参考数据准备失败：{e}")
        return

    print(f"读取文件：{input_xlsx}")
    xls = pd.ExcelFile(input_xlsx)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    headers = [
        "Chromosome", "Position", "Strand",
        "DNA_short_raw",
        "Mutation_Positions",
        "Ref_600bp_from_FASTA",
        "Pair_ID", "Pair_Status", "Non_Perfect",
        "Forward_Primer", "Forward_Tm", "Forward_Specificity",
        "Reverse_Primer", "Reverse_Tm", "Reverse_Specificity",
        "Amplicon_Seq", "Amplicon_Size",
        "Fwd_Offset_bp", "Rev_Offset_bp", "R_Offset_Pref"
    ]

    cpu = os.cpu_count() or 4
    jobs = args.jobs if args.jobs and args.jobs > 0 else max(2, cpu // 2)
    jobs = min(jobs, max(2, cpu))
    if args.blast_threads and args.blast_threads > 0:
        per_threads = args.blast_threads
    else:
        per_threads = max(1, (cpu - 1) // jobs)

    print(f"[并行] 目标工人: {jobs} ；每任务 BLAST 线程: {per_threads} ；提交分块: {args.chunk}")

    for sheet in xls.sheet_names:
        print(f"\n---- 处理 sheet: {sheet} ----")
        try:
            hdr = detect_header_row(input_xlsx, sheet)
            if hdr is not None:
                df = pd.read_excel(input_xlsx, sheet_name=sheet, header=hdr, dtype=str)
            else:
                df = pd.read_excel(input_xlsx, sheet_name=sheet, dtype=str)
        except Exception as e:
            print(f"  读取 sheet {sheet} 失败: {e}")
            continue

        df = df.dropna(axis=1, how='all').reset_index(drop=True)
        if df.empty:
            print(f"  Sheet {sheet} 为空，跳过")
            continue

        # 查找DNA序列列
        target_col = find_col(df, ['dna', 'sequence', 'target', 'crrna', 'bulge', 'seq'])
        
        if not target_col:
            print(f"  警告: sheet {sheet} 未找到DNA序列列")
            continue

        # 构建任务列表
        tasks = []
        for i, row in df.iterrows():
            dna_short = extract_dna_from_row(df, i, target_col, look_ahead=3) or ""
            if dna_short:
                tasks.append((sheet, dna_short, base_dir, fasta_name, db_prefix, per_threads))
            else:
                print(f"  行 {i+1} 未找到有效DNA序列")

        if not tasks:
            print(f"  Sheet {sheet} 没有有效任务")
            continue

        print(f"  找到 {len(tasks)} 个有效序列")

        # 选择执行器（带回退）
        executor_type = "thread"
        ex = None
        if args.force_executor in ("auto", "process"):
            try_jobs = jobs
            while try_jobs >= 2:
                try:
                    print(f"  尝试创建进程池 ({try_jobs} 进程)...")
                    ex = _try_make_process_pool(try_jobs)
                    executor_type = "process"
                    if try_jobs != jobs:
                        print(f"[回退成功] 以 {try_jobs} 个进程运行（原计划 {jobs}）。")
                    break
                except OSError as e:
                    # WinError 1455 或其他 OSError
                    print(f"[警告] 创建 {try_jobs} 进程失败：{e}. 尝试减少进程数...")
                    try_jobs = max(2, try_jobs // 2)
                except Exception as e:
                    print(f"[警告] 进程池启动异常（{type(e).__name__}: {e}）。尝试减少进程数...")
                    try_jobs = max(2, try_jobs // 2)
        if ex is None:
            if args.force_executor == "process":
                print("[警告] 强制 process 但无法创建，自动回退为 thread。")
            ex = ThreadPoolExecutor(max_workers=jobs)
            executor_type = "thread"
        print(f"[执行器] 使用 {executor_type} 池。")

        # 分批提交任务，降低一次性内存压力
        results = []
        total_tasks = len(tasks)
        processed = 0
        
        def submit_chunk(chunk):
            futs = [ex.submit(worker_process, t) for t in chunk]
            for fut in as_completed(futs):
                results.append(fut.result())

        if args.chunk and args.chunk > 0:
            for i in range(0, total_tasks, args.chunk):
                chunk = tasks[i:i + args.chunk]
                processed += len(chunk)
                print(f"  处理块 {i//args.chunk + 1}/{(total_tasks-1)//args.chunk + 1} ({len(chunk)} 条序列, 进度 {processed}/{total_tasks})")
                submit_chunk(chunk)
        else:
            submit_chunk(tasks)

        # 用完就关闭执行器（释放内存）
        ex.shutdown(wait=True, cancel_futures=False)

        # 写到 sheet
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        for s, outrow in results:
            ws.append(outrow)

        wb.save(output_xlsx)
        print(f"  [保存] {sheet} 已写入 {len(results)} 行")

    wb.save(output_xlsx)
    print(f"\n主结果保存到: {output_xlsx}")
    print("完成。")

if __name__ == "__main__":
    main()
